#!/usr/bin/env python3
"""Build a Google-Sheets-ready family budget workbook from a config JSON.

Usage:
  python build_workbook.py --config config.json --out family_budget.xlsx
  python build_workbook.py --out demo.xlsx        # builds a demo if no config

Design notes:
- All dashboard cells are live formulas (SUMIFS over the Transactions tab) so the
  sheet stays correct as rows are added.
- Dropdowns use data validation that survives the Excel -> Google Sheets import.
- Hebrew sheets are set right-to-left.
- Every division is IFERROR-guarded (no #DIV/0!).
"""
import argparse, json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ── Canonical taxonomy: Hebrew group -> (type, exclude_default, [hebrew leaves])
TAXONOMY = {
    "הכנסות": ("income", False, ["משכורת", "פרילנס", "שכר דירה", "דיבידנדים",
        "החזר מס", "קצבאות", "מתנות שהתקבלו", "הכנסות אחרות"]),
    "מזון וסופרמרקט": ("expense", False, ["סופרמרקט", "שוק", "מסעדות", "קפה",
        "משלוחים", "בר ואלכוהול"]),
    "דיור": ("expense", False, ["שכירות", "משכנתא", "ארנונה", "ועד בית", "חשמל",
        "גז", "מים", "אינטרנט", "טלפון", "ביטוח דירה", "תחזוקת בית"]),
    "תחבורה": ("expense", False, ["דלק", "ביטוח רכב", "רישוי וטסט", "חנייה",
        "כביש 6", "תחבורה ציבורית", "מוניות", "תחזוקת רכב", "קנסות תנועה"]),
    "בריאות": ("expense", False, ["קופת חולים", "ביטוח בריאות", "תרופות",
        "רופאים", "שיניים", "אופטיקה"]),
    "ביטוחים": ("expense", False, ["ביטוח חיים", "ביטוח מנהלים", "ביטוח נסיעות",
        "ביטוח אחר"]),
    "חינוך": ("expense", False, ["שכר לימוד", "קורסים", "ספרים וחומרי לימוד",
        "גן / מעון", "חוגים"]),
    "ילדים": ("expense", False, ["מזון לתינוקות", "ביגוד ילדים", "צעצועים",
        "בייביסיטר"]),
    "בידור": ("expense", False, ["סטרימינג", "מוזיקה", "משחקים",
        "קולנוע ואירועים", "ספורט וחדר כושר", "ספרים ומגזינים"]),
    "קניות": ("expense", False, ["ביגוד והנעלה", "אלקטרוניקה", "ריהוט ובית",
        "מתנות", "קניות אונליין"]),
    "מסעות": ("expense", False, ["טיסות", "מלונות", "השכרת רכב", "אטרקציות"]),
    "מנויים": ("expense", False, ["מנוי תוכן", "מנוי תוכנה", "מנוי חדר כושר",
        "מנוי אחר"]),
    "חיסכון והשקעות": ("investment", True, ["קרן פנסיה", "קרן השתלמות",
        "קופת גמל", "מניות", "חיסכון", "קריפטו"]),
    "הלוואות וחובות": ("expense", False, ["החזר הלוואה", "כרטיס אשראי", "ריבית"]),
    "עסק": ("expense", False, ["הוצאות עסקיות", "ציוד משרדי", "שירותים מקצועיים",
        "פרסום"]),
    "העברות": ("transfer", True, ["ביט", "פייבוקס", "העברה בנקאית",
        "משיכת מזומן"]),
    "מיסים ועמלות": ("expense", False, ["מס הכנסה", "ביטוח לאומי", "עמלות בנק",
        "עמלות כרטיס"]),
    "לא מסווג": ("expense", False, ["ממתין לסיווג", "אחר"]),
}
EXPENSE_GROUPS = [g for g, (t, _, _) in TAXONOMY.items() if t == "expense"]

# ── Styling helpers
NAVY = "1F3A5F"; TEAL = "2E8B8B"; LIGHT = "EAF2F2"; AMBER = "F4B400"
HEAD = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE = Font(name="Arial", bold=True, color=NAVY, size=16)
SUB = Font(name="Arial", bold=True, color=NAVY, size=12)
BASE = Font(name="Arial", size=11)
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
SUBHEAD_FILL = PatternFill("solid", fgColor=TEAL)
LIGHT_FILL = PatternFill("solid", fgColor=LIGHT)
AMBER_FILL = PatternFill("solid", fgColor=AMBER)
THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hdr(ws, row, headers, start=1, fill=HEAD_FILL):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start + i, value=h)
        c.font = HEAD; c.fill = fill; c.alignment = Alignment("center", "center")
        c.border = BORDER


def money_fmt(sym):
    return f'#,##0" {sym}";-#,##0" {sym}";"-"'


def build(cfg):
    sym = cfg.get("currency_symbol", "₪")
    rtl = cfg.get("rtl", True)
    M = money_fmt(sym)
    wb = Workbook()

    # ════════ Lists (dropdown sources) ════════
    lst = wb.active; lst.title = "Lists"
    lst.sheet_view.rightToLeft = rtl
    lst["A1"] = "Types"; lst["B1"] = "Groups"; lst["C1"] = "Categories"
    lst["D1"] = "Members"; lst["E1"] = "Accounts"
    for c in "ABCDE": lst[f"{c}1"].font = HEAD; lst[f"{c}1"].fill = HEAD_FILL
    for i, t in enumerate(["income", "expense", "transfer", "investment"]):
        lst.cell(row=2 + i, column=1, value=t)
    groups = list(TAXONOMY.keys())
    for i, g in enumerate(groups):
        lst.cell(row=2 + i, column=2, value=g)
    cats = [leaf for (_, _, leaves) in TAXONOMY.values() for leaf in leaves]
    for i, c in enumerate(cats):
        lst.cell(row=2 + i, column=3, value=c)
    members = cfg.get("members", []) or ["—"]
    for i, m in enumerate(members):
        lst.cell(row=2 + i, column=4, value=m)
    accounts = cfg.get("accounts", []) or ["—"]
    for i, a in enumerate(accounts):
        lst.cell(row=2 + i, column=5, value=a)
    for col, w in zip("ABCDE", [12, 20, 22, 16, 18]):
        lst.column_dimensions[col].width = w
    lst.sheet_state = "hidden"
    g_end = 1 + len(groups); c_end = 1 + len(cats)
    m_end = 1 + len(members); a_end = 1 + len(accounts)

    # ════════ README ════════
    rm = wb.create_sheet("הוראות")
    rm.sheet_view.rightToLeft = rtl
    rm.sheet_view.showGridLines = False
    rm["A1"] = "מתכנן תקציב משפחתי · Family Budget Planner"; rm["A1"].font = TITLE
    lines = [
        ("", SUB),
        ("איך מתחילים — 3 צעדים / Get started in 3 steps", SUB),
        ("1. העלו את הקובץ ל-Google Drive  ·  Upload this file to Google Drive", BASE),
        ("2. פתחו עם Google Sheets  ·  Right-click → Open with Google Sheets", BASE),
        ("3. שתפו עם המשפחה (כפתור Share)  ·  Share with the family", BASE),
        ("", BASE),
        ("איך מעדכנים הוצאות / How to log spending", SUB),
        ("פשוט כתבו ל-Claude בשפה חופשית, למשל:", BASE),
        ('   « שילמתי 250 בשופרסל וקנס חניה 100 »', BASE),
        ('   « salary 18,500 came in, moved 2,000 to keren hishtalmut »', BASE),
        ("Claude יחזיר שורות מסווגות להדבקה בלשונית « תנועות ».", BASE),
        ("", BASE),
        ("הלשוניות / The tabs", SUB),
        ("• הגדרות — Setup: family, cycle, salary day, currency", BASE),
        ("• תקציב — Budget: monthly target per category group", BASE),
        ("• תנועות — Transactions: the ledger (where rows get pasted)", BASE),
        ("• לוח בקרה — Dashboard: income vs spend, savings rate, per-group", BASE),
        ("• יעדים — Goals: savings/payoff goals, required ₪/month, on-track", BASE),
        ("• המלצות — Recommendations: auto-insights + savings action board", BASE),
        ("• נכסים — Net Worth: assets, liabilities, net position", BASE),
        ("", BASE),
        ("הערה: התקציב השוטף לא כולל השקעות והעברות (תנועות נכסים).", BASE),
        ("Note: the operating budget excludes investments & transfers.", BASE),
    ]
    for i, (txt, f) in enumerate(lines, start=2):
        rm[f"A{i}"] = txt; rm[f"A{i}"].font = f
    rm.column_dimensions["A"].width = 78

    # ════════ Setup ════════
    st = wb.create_sheet("הגדרות")
    st.sheet_view.rightToLeft = rtl
    st["A1"] = "הגדרות · Setup"; st["A1"].font = TITLE
    rows = [
        ("שם משק הבית / Household", cfg.get("household", "")),
        ("מדינה / Country", cfg.get("country", "ישראל")),
        ("מטבע / Currency", sym),
        ("יום קבלת משכורת / Salary day", cfg.get("salary_day", 10)),
        ("תחילת מחזור (יום) / Cycle start day", cfg.get("cycle_start_day", 15)),
    ]
    r = 3
    for label, val in rows:
        st.cell(row=r, column=1, value=label).font = SUB
        c = st.cell(row=r, column=2, value=val); c.font = BASE
        c.fill = LIGHT_FILL; c.border = BORDER
        r += 1
    CYCLE_DAY = "הגדרות!$B$7"  # cycle start day cell
    st.cell(row=r + 1, column=1, value="חברי משפחה / Members").font = SUB
    for i, m in enumerate(members):
        st.cell(row=r + 2 + i, column=1, value=m).font = BASE
    ar = r + 2 + len(members) + 1
    st.cell(row=ar, column=1, value="חשבונות וכרטיסים / Accounts & cards").font = SUB
    for i, a in enumerate(accounts):
        st.cell(row=ar + 1 + i, column=1, value=a).font = BASE
    st.column_dimensions["A"].width = 34; st.column_dimensions["B"].width = 22

    # ════════ Budget targets ════════
    bd = wb.create_sheet("תקציב")
    bd.sheet_view.rightToLeft = rtl
    bd["A1"] = "תקציב חודשי · Monthly Budget"; bd["A1"].font = TITLE
    hdr(bd, 3, ["קבוצה / Group", "יעד חודשי / Monthly target"])
    targets = cfg.get("targets", {})
    for i, g in enumerate(EXPENSE_GROUPS):
        rr = 4 + i
        bd.cell(row=rr, column=1, value=g).font = BASE
        c = bd.cell(row=rr, column=2, value=targets.get(g, 0))
        c.font = Font(name="Arial", size=11, color="0000FF")  # input = blue
        c.number_format = M; c.fill = LIGHT_FILL; c.border = BORDER
    tot_row = 4 + len(EXPENSE_GROUPS)
    bd.cell(row=tot_row, column=1, value="סה\"כ / Total").font = SUB
    tc = bd.cell(row=tot_row, column=2,
                 value=f"=SUM(B4:B{tot_row - 1})")
    tc.number_format = M; tc.font = SUB
    bd.column_dimensions["A"].width = 24; bd.column_dimensions["B"].width = 20

    # ════════ Transactions ════════
    tx = wb.create_sheet("תנועות")
    tx.sheet_view.rightToLeft = rtl
    cols = ["תאריך/Date", "בן משפחה/Member", "חשבון/Account", "סוג/Type",
            "קבוצה/Group", "קטגוריה/Category", "תיאור/Description",
            "סכום/Amount", "מחוץ לתקציב/Exclude", "הערות/Notes"]
    hdr(tx, 1, cols)
    sample = [
        ("=TODAY()", members[0], accounts[0], "income", "הכנסות", "משכורת",
         "משכורת נטו", cfg.get("sample_salary", 18500), False, ""),
        ("=TODAY()", members[0], accounts[-1], "expense", "מזון וסופרמרקט",
         "סופרמרקט", "שופרסל", 420, False, "דוגמה"),
        ("=TODAY()", members[0], accounts[0], "investment", "חיסכון והשקעות",
         "קרן השתלמות", "הפקדה", 2000, True, "לא נספר בתקציב השוטף"),
    ]
    for i, row in enumerate(sample):
        rr = 2 + i
        for j, val in enumerate(row):
            c = tx.cell(row=rr, column=1 + j, value=val)
            c.font = BASE; c.border = BORDER
            if j == 0: c.number_format = "yyyy-mm-dd"
            if j == 7: c.number_format = M
    widths = [13, 15, 14, 12, 18, 18, 22, 13, 14, 20]
    for i, w in enumerate(widths):
        tx.column_dimensions[get_column_letter(1 + i)].width = w
    tx.freeze_panes = "A2"
    LAST = 1000

    def dv(formula, col):
        d = DataValidation(type="list", formula1=formula, allow_blank=True)
        tx.add_data_validation(d)
        d.add(f"{col}2:{col}{LAST}")
    dv(f"Lists!$D$2:$D${m_end}", "B")
    dv(f"Lists!$E$2:$E${a_end}", "C")
    dv(f"Lists!$A$2:$A$5", "D")
    dv(f"Lists!$B$2:$B${g_end}", "E")
    dv(f"Lists!$C$2:$C${c_end}", "F")
    dE = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    tx.add_data_validation(dE); dE.add(f"I2:I{LAST}")

    # ════════ Dashboard ════════
    db = wb.create_sheet("לוח בקרה")
    db.sheet_view.rightToLeft = rtl
    db.sheet_view.showGridLines = False
    db["A1"] = "לוח בקרה · Dashboard"; db["A1"].font = TITLE
    # cycle window helper cells
    db["A3"] = "מחזור נוכחי / Current cycle"; db["A3"].font = SUB
    db["A4"] = "מתאריך / From"; db["A4"].font = BASE
    db["B4"] = (f"=IF(DAY(TODAY())>={CYCLE_DAY},"
                f"DATE(YEAR(TODAY()),MONTH(TODAY()),{CYCLE_DAY}),"
                f"EDATE(DATE(YEAR(TODAY()),MONTH(TODAY()),{CYCLE_DAY}),-1))")
    db["A5"] = "עד / To"; db["A5"].font = BASE
    db["B5"] = "=EDATE(B4,1)-1"
    for cc in ("B4", "B5"):
        db[cc].number_format = "yyyy-mm-dd"; db[cc].font = BASE
        db[cc].fill = LIGHT_FILL; db[cc].border = BORDER
    S = "$B$4"; E = "$B$5"
    Tx = "תנועות!"
    rng_amt = f"{Tx}$H:$H"; rng_type = f"{Tx}$D:$D"
    rng_date = f"{Tx}$A:$A"; rng_excl = f"{Tx}$I:$I"; rng_grp = f"{Tx}$E:$E"

    def cycle_sumifs(type_val, extra=""):
        return (f'=SUMIFS({rng_amt},{rng_type},"{type_val}",'
                f'{rng_date},">="&{S},{rng_date},"<="&{E}{extra})')

    kpis = [
        ("הכנסות / Income", cycle_sumifs("income")),
        ("הוצאות שוטפות / Operating spend",
         f'={cycle_sumifs("expense")[1:]}-SUMIFS({rng_amt},{rng_type},'
         f'"expense",{rng_excl},TRUE,{rng_date},">="&{S},{rng_date},'
         f'"<="&{E})'),
        ("נטו שוטף / Net operating", "=B8-B9"),
        ("שיעור חיסכון / Savings rate", '=IFERROR(B10/B8,"-")'),
        ("הושקע החודש / Invested", cycle_sumifs("investment")),
        ("תזרים יוצא כולל / Total outflow", "=B9+B12"),
    ]
    rr = 8
    for label, formula in kpis:
        db.cell(row=rr, column=1, value=label).font = SUB
        c = db.cell(row=rr, column=2, value=formula)
        c.font = BASE; c.fill = LIGHT_FILL; c.border = BORDER
        c.number_format = "0.0%" if "Savings" in label or "חיסכון" in label and rr == 11 else M
        rr += 1
    db["B11"].number_format = "0.0%"
    # per-group table
    gr = rr + 1
    db.cell(row=gr, column=1, value="לפי קבוצה / By group").font = SUB
    hdr(db, gr + 1, ["קבוצה / Group", "יעד / Target", "בפועל / Actual",
                     "פער / Variance"])
    base = gr + 2
    for i, g in enumerate(EXPENSE_GROUPS):
        row = base + i
        db.cell(row=row, column=1, value=g).font = BASE
        db.cell(row=row, column=2,
                value=f'=IFERROR(VLOOKUP(A{row},תקציב!$A:$B,2,FALSE),0)'
                ).number_format = M
        db.cell(row=row, column=3,
                value=(f'=SUMIFS({rng_amt},{rng_grp},A{row},{rng_type},'
                       f'"expense",{rng_date},">="&{S},{rng_date},"<="&{E})')
                ).number_format = M
        vc = db.cell(row=row, column=4, value=f"=B{row}-C{row}")
        vc.number_format = M
        for col in range(1, 5):
            db.cell(row=row, column=col).border = BORDER
            db.cell(row=row, column=col).font = BASE
    db.column_dimensions["A"].width = 28
    for col in "BCD":
        db.column_dimensions[col].width = 16
    pg_first = base
    pg_last = base + len(EXPENSE_GROUPS) - 1
    DB = "'לוח בקרה'!"
    BLUE = Font(name="Arial", color="0000FF")

    # ════════ Goals ════════
    gl = wb.create_sheet("יעדים")
    gl.sheet_view.rightToLeft = rtl
    gl["A1"] = "יעדים פיננסיים · Financial Goals"; gl["A1"].font = TITLE
    gl["A2"] = "מלאו יעד, תאריך וכמה נצבר — הגיליון מחשב כמה צריך להפריש לחודש."
    gl["A2"].font = BASE
    hdr(gl, 3, ["יעד / Goal", "סוג / Type", "יעד סכום / Target",
                "תאריך יעד / Target date", "נצבר / Saved",
                "חודשים נותרו / Months", "עוד צריך / Still needed",
                "נדרש לחודש / Per month", "התקדמות / Progress"])
    goals = cfg.get("goals") or [
        {"name": "קרן חירום / Emergency fund", "type": "חיסכון",
         "target": 60000, "date": "2027-06-30", "saved": 0},
        {"name": "טיול / Trip", "type": "חיסכון",
         "target": 25000, "date": "2026-07-22", "saved": 0}]
    r0 = 4
    for i, gobj in enumerate(goals):
        row = r0 + i
        gl.cell(row=row, column=1, value=gobj.get("name", "")).font = BASE
        gl.cell(row=row, column=2, value=gobj.get("type", "חיסכון")).font = BASE
        for col, key, fmt in [(3, "target", M), (4, "date", "yyyy-mm-dd"),
                              (5, "saved", M)]:
            c = gl.cell(row=row, column=col, value=gobj.get(key, 0))
            c.number_format = fmt; c.font = BLUE; c.fill = LIGHT_FILL
        gl.cell(row=row, column=6, value=(
            f"=IFERROR(MAX(0,(YEAR(D{row})-YEAR(TODAY()))*12+"
            f"(MONTH(D{row})-MONTH(TODAY()))),0)")).font = BASE
        gl.cell(row=row, column=7,
                value=f"=MAX(0,C{row}-E{row})").number_format = M
        gl.cell(row=row, column=8,
                value=f"=IF(F{row}<=0,G{row},G{row}/F{row})").number_format = M
        p = gl.cell(row=row, column=9, value=f"=IFERROR(E{row}/C{row},0)")
        p.number_format = "0%"
        for col in range(1, 10):
            cc = gl.cell(row=row, column=col)
            cc.border = BORDER
            if cc.font.color is None or cc.font.name != "Arial":
                cc.font = BASE
        gl.cell(row=row, column=7).font = BASE
        gl.cell(row=row, column=8).font = BASE
    glast = r0 + len(goals) - 1
    tr = glast + 1
    gl.cell(row=tr, column=7, value="נדרש לחודש סה\"כ →").font = SUB
    tc = gl.cell(row=tr, column=8, value=f"=SUM(H{r0}:H{glast})")
    tc.number_format = M; tc.font = SUB
    gl.cell(row=tr + 1, column=7, value="עודף חודשי / Surplus").font = BASE
    sc = gl.cell(row=tr + 1, column=8, value=f"={DB}$B$10")
    sc.number_format = M; sc.font = BASE
    gl.cell(row=tr + 2, column=7, value="פער / Gap").font = SUB
    gp = gl.cell(row=tr + 2, column=8, value=f"={DB}$B$10-H{tr}")
    gp.number_format = M; gp.font = SUB; gp.fill = AMBER_FILL
    dvt = DataValidation(type="list", formula1='"חיסכון,החזר חוב"',
                         allow_blank=True)
    gl.add_data_validation(dvt); dvt.add(f"B{r0}:B{r0 + 40}")
    for col, w in zip("ABCDEFGHI", [24, 12, 13, 14, 13, 11, 13, 13, 11]):
        gl.column_dimensions[col].width = w

    # ════════ Recommendations ════════
    rc = wb.create_sheet("המלצות")
    rc.sheet_view.rightToLeft = rtl
    rc.sheet_view.showGridLines = False
    rc["A1"] = "המלצות חיסכון · Savings Recommendations"; rc["A1"].font = TITLE
    rc["A3"] = "תובנות אוטומטיות / Auto-insights"; rc["A3"].font = SUB
    grp_rng = f"{DB}$A${pg_first}:$A${pg_last}"
    act_rng = f"{DB}$C${pg_first}:$C${pg_last}"
    var_rng = f"{DB}$D${pg_first}:$D${pg_last}"
    insights = [
        ("הקטגוריה הגדולה / Biggest category",
         f'=IFERROR(INDEX({grp_rng},MATCH(MAX({act_rng}),{act_rng},0)),"-")',
         None),
        ("   ↳ סכום / Amount", f"=MAX({act_rng})", M),
        ("הכי חורג / Most over budget",
         f'=IFERROR(INDEX({grp_rng},MATCH(MIN({var_rng}),{var_rng},0)),"-")',
         None),
        ("   ↳ חריגה / Over by", f"=MAX(0,-MIN({var_rng}))", M),
        ("מנויים החודש / Subscriptions",
         f'=SUMIFS(תנועות!$H:$H,תנועות!$E:$E,"מנויים",תנועות!$A:$A,'
         f'">="&{DB}$B$4,תנועות!$A:$A,"<="&{DB}$B$5)', M),
        ("שיעור חיסכון נוכחי / Savings rate", f"={DB}$B$11", "0.0%"),
    ]
    r = 4
    for label, formula, fmt in insights:
        rc.cell(row=r, column=1, value=label).font = BASE
        c = rc.cell(row=r, column=2, value=formula)
        c.font = BASE; c.fill = LIGHT_FILL; c.border = BORDER
        if fmt:
            c.number_format = fmt
        r += 1
    ab = r + 1
    rc.cell(row=ab, column=1,
            value="לוח פעולות / Action board — Claude ממלא לפי ההוצאות והמועדונים"
            ).font = SUB
    hdr(rc, ab + 1, ["תחום / Area", "פעולה / Action (מועדון/מהלך)",
                     "חיסכון ₪/חודש / Est. saving", "סטטוס / Status"])
    seed = [
        ("מזון / Groceries", "מועדון סופר + cashback באשראי", 0, "לביצוע"),
        ("אוכל בחוץ / Dining", "ניצול תן ביס / Cibus", 0, "לביצוע"),
        ("מנויים / Subscriptions", "ביטול מנויים כפולים", 0, "לביצוע"),
    ]
    sr = ab + 2
    for i, (area, action, save, status) in enumerate(seed):
        row = sr + i
        rc.cell(row=row, column=1, value=area).font = BASE
        rc.cell(row=row, column=2, value=action).font = BASE
        cs = rc.cell(row=row, column=3, value=save)
        cs.number_format = M; cs.font = BLUE; cs.fill = LIGHT_FILL
        rc.cell(row=row, column=4, value=status).font = BASE
        for col in range(1, 5):
            rc.cell(row=row, column=col).border = BORDER
    LASTR = sr + 40
    dvs = DataValidation(type="list", formula1='"לביצוע,בוצע"',
                         allow_blank=True)
    rc.add_data_validation(dvs); dvs.add(f"D{sr}:D{LASTR}")
    tot = LASTR + 2
    rc.cell(row=tot, column=2, value="חיסכון פוטנציאלי (לביצוע) →").font = SUB
    pot = rc.cell(row=tot, column=3,
                  value=f'=SUMIF(D{sr}:D{LASTR},"לביצוע",C{sr}:C{LASTR})')
    pot.number_format = M; pot.font = SUB
    rc.cell(row=tot + 1, column=2,
            value="שיעור חיסכון משופר / Improved rate").font = BASE
    imp = rc.cell(row=tot + 1, column=3,
                  value=f'=IFERROR(({DB}$B$10+C{tot})/{DB}$B$8,"-")')
    imp.number_format = "0.0%"; imp.font = BASE
    for col, w in zip("ABCD", [28, 34, 18, 14]):
        rc.column_dimensions[col].width = w

    # ════════ Net Worth ════════
    nw = wb.create_sheet("נכסים")
    nw.sheet_view.rightToLeft = rtl
    nw["A1"] = "נכסים והון עצמי · Net Worth"; nw["A1"].font = TITLE
    hdr(nw, 3, ["נכס / Asset", "ערך / Value"])
    asset_rows = cfg.get("assets") or [
        ["עו\"ש / Checking", 0], ["חיסכון / Savings", 0],
        ["קרן השתלמות", 0], ["קרן פנסיה", 0], ["קופת גמל", 0],
        ["תיק מניות / Brokerage", 0]]
    r = 4
    for name, val in asset_rows:
        nw.cell(row=r, column=1, value=name).font = BASE
        c = nw.cell(row=r, column=2, value=val)
        c.number_format = M; c.font = Font(name="Arial", color="0000FF")
        c.fill = LIGHT_FILL; c.border = BORDER
        r += 1
    a_first, a_last = 4, r - 1
    nw.cell(row=r, column=1, value="סה\"כ נכסים / Total assets").font = SUB
    nw.cell(row=r, column=2,
            value=f"=SUM(B{a_first}:B{a_last})").number_format = M
    nw.cell(row=r, column=2).font = SUB
    lr = r + 2
    hdr(nw, lr, ["התחייבות / Liability", "ערך / Value"])
    liab_rows = cfg.get("liabilities") or [
        ["משכנתא / Mortgage", 0], ["הלוואות / Loans", 0],
        ["יתרת כרטיסים / Card balances", 0]]
    r = lr + 1
    for name, val in liab_rows:
        nw.cell(row=r, column=1, value=name).font = BASE
        c = nw.cell(row=r, column=2, value=val)
        c.number_format = M; c.font = Font(name="Arial", color="0000FF")
        c.fill = LIGHT_FILL; c.border = BORDER
        r += 1
    l_first, l_last = lr + 1, r - 1
    nw.cell(row=r, column=1, value="סה\"כ התחייבויות / Total").font = SUB
    nw.cell(row=r, column=2,
            value=f"=SUM(B{l_first}:B{l_last})").number_format = M
    nw.cell(row=r, column=2).font = SUB
    nwr = r + 2
    nw.cell(row=nwr, column=1, value="הון עצמי נטו / Net worth").font = TITLE
    c = nw.cell(row=nwr, column=2, value=f"=B{a_last + 1}-B{r}")
    c.number_format = M; c.font = TITLE; c.fill = AMBER_FILL; c.border = BORDER
    nw.column_dimensions["A"].width = 30; nw.column_dimensions["B"].width = 18

    # order tabs
    wb.move_sheet("הוראות", -(wb.index(wb["הוראות"])))
    return wb


def demo_config():
    return {
        "household": "משפחת קיידנוב / Kaidanov Family",
        "country": "ישראל", "currency_symbol": "₪", "rtl": True,
        "salary_day": 10, "cycle_start_day": 15,
        "members": ["Gregory", "בן/בת זוג", "ילד/ה"],
        "accounts": ["בנק עו\"ש", "ויזה", "מקס/Max", "מזומן"],
        "sample_salary": 18500,
        "targets": {"מזון וסופרמרקט": 4000, "דיור": 6500, "תחבורה": 2500,
                    "בריאות": 800, "חינוך": 1500, "בידור": 900,
                    "קניות": 1200, "מנויים": 350},
        "assets": [["עו\"ש", 0], ["קרן השתלמות", 0], ["קרן פנסיה", 0],
                   ["קופת גמל", 0], ["תיק מניות", 0]],
        "liabilities": [["משכנתא", 0], ["הלוואות", 0]],
        "goals": [
            {"name": "טיול לאיטליה / Italy trip", "type": "חיסכון",
             "target": 30000, "date": "2026-07-22", "saved": 12000},
            {"name": "קרן חירום / Emergency fund", "type": "חיסכון",
             "target": 60000, "date": "2027-06-30", "saved": 15000},
            {"name": "החזר הלוואה / Loan payoff", "type": "החזר חוב",
             "target": 40000, "date": "2027-12-31", "saved": 8000}],
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config")
    ap.add_argument("--out", required=True)
    a = ap.parse_args()
    if a.config:
        with open(a.config, encoding="utf-8") as f:
            cfg = json.load(f)
    else:
        cfg = demo_config()
        print("No --config given; building demo workbook.", file=sys.stderr)
    wb = build(cfg)
    wb.save(a.out)
    print(f"Saved {a.out}")


if __name__ == "__main__":
    main()
