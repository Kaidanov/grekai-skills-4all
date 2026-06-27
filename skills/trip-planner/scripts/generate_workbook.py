"""
generate_workbook.py — render a generic trip bundle into a 7-sheet .xlsx
(Itinerary, Attractions, Foodies, Budget, Costs Remarks, Notes and tips, Prompt).

Not specific to any destination. Pass a bundle dict (see assets/bundle_schema.json).
Budget uses live formulas. Attractions/Foodies get clickable source links.

Usage:  python generate_workbook.py bundle.json out.xlsx
"""
from __future__ import annotations
import json, sys
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HF = PatternFill("solid", fgColor="1F4E5F"); HFONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUB = PatternFill("solid", fgColor="DDEBF0"); SEC = Font(name="Arial", size=11, bold=True, color="1F4E5F")
STAR = PatternFill("solid", fgColor="FFF6D5")
BASE = Font(name="Arial", size=10); BOLD = Font(name="Arial", size=10, bold=True)
LINK = Font(name="Arial", size=10, color="1155CC", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top"); TOP = Alignment(vertical="top")
CTR = Alignment(wrap_text=True, vertical="center", horizontal="center")
thin = Side(style="thin", color="BBBBBB"); BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
EUR = "#,##0.00"


def _gm(n, c): return f"https://www.google.com/maps/search/?api=1&query={quote(f'{n} {c}')}"
def _ta(n, c): return f"https://www.tripadvisor.com/Search?q={quote(f'{n} {c}')}"
def _gr(n, c): return f"https://www.google.com/search?q={quote(f'{n} {c} reviews')}"


def _hdr(ws, row, n):
    for c in range(1, n + 1):
        x = ws.cell(row=row, column=c); x.fill = HF; x.font = HFONT; x.alignment = CTR; x.border = BORDER


def _link(ws, r, c, url):
    x = ws.cell(row=r, column=c, value=url); x.hyperlink = url; x.font = LINK; x.alignment = TOP; x.border = BORDER


def _links(item, name, city):
    lk = item.get("links") or {}
    return (lk.get("maps") or _gm(name, city),
            lk.get("tripadvisor") or _ta(name, city),
            lk.get("google_reviews") or _gr(name, city))


def build(bundle: dict, out: str):
    wb = Workbook()
    cur = (bundle.get("trip", {}).get("currency") or "").strip()
    money = (f'"{_sym(cur)}"' + EUR) if cur else EUR

    # 1. Itinerary
    ws = wb.active; ws.title = "Itenarary"
    t = bundle.get("trip", {})
    ws["A1"] = t.get("name", "Trip"); ws["A1"].font = SEC
    ws["A2"] = f"{t.get('date_start','')} → {t.get('date_end','')}  ·  {_travelers(t)}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True)
    ws.append([]); h = ["Date", "Day", "Location", "Activities", "Accommodation", "Notes"]; ws.append(h)
    hr = ws.max_row; _hdr(ws, hr, 6)
    for r in bundle.get("itinerary", []):
        ws.append([r.get("date"), r.get("day"), r.get("location"),
                   r.get("activities"), r.get("accommodation"), r.get("notes", "")])
    for row in ws.iter_rows(min_row=hr + 1, max_row=ws.max_row):
        for x in row: x.font = BASE; x.alignment = WRAP; x.border = BORDER
    for i, w in enumerate([11, 6, 26, 50, 28, 40], 1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)

    # 2. Attractions
    ws = wb.create_sheet("Attractions")
    h = ["Name", "Type", "Location", "Highlights", "Accessibility", "Cost", "Popularity Grade",
         "Traveler's Reviews Grade", "Suitable For", "Special Features", "Interesting Reviews / Notes",
         "Google Maps", "Tripadvisor", "Google Reviews"]
    ws.append(h); _hdr(ws, 1, len(h)); r = 2
    for a in bundle.get("attractions", []):
        vals = [a.get(k, "") for k in ("name", "type", "location", "highlights", "accessibility",
                "cost", "popularity", "reviews", "suitable_for", "special", "notes")]
        for ci, v in enumerate(vals, 1): ws.cell(row=r, column=ci, value=v)
        for ci, u in zip((12, 13, 14), _links(a, a.get("name", ""), a.get("location", ""))): _link(ws, r, ci, u)
        if a.get("importer_flag") or "⭐" in str(a.get("special", "")):
            for ci in range(1, 12): ws.cell(row=r, column=ci).fill = STAR
        r += 1
    _grid(ws, 2, 14)
    for i, w in enumerate([30, 20, 20, 34, 20, 18, 12, 12, 15, 26, 40, 15, 15, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # 3. Foodies
    ws = wb.create_sheet("Foodies")
    h = ["Name of the Place", "Average Price", "Location", "Popularity Grade", "Type of Dishes",
         "Specialty", "Ambiance Description", "Interesting Reviews / Notes",
         "Google Maps", "Tripadvisor", "Google Reviews"]
    ws.append(h); _hdr(ws, 1, len(h)); r = 2
    for f in bundle.get("foodies", []):
        vals = [f.get(k, "") for k in ("name", "price", "location", "popularity",
                "dishes", "specialty", "ambiance", "notes")]
        for ci, v in enumerate(vals, 1): ws.cell(row=r, column=ci, value=v)
        for ci, u in zip((9, 10, 11), _links(f, f.get("name", ""), f.get("location", ""))): _link(ws, r, ci, u)
        r += 1
    _grid(ws, 2, 11)
    for i, w in enumerate([26, 14, 22, 12, 20, 26, 26, 40, 15, 15, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # 4. Budget (live formulas)
    ws = wb.create_sheet("Budget"); R = 1
    def put(vals, font=BASE, fill=None):
        nonlocal R
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=R, column=i, value=v); c.font = font; c.alignment = WRAP
            if fill: c.fill = fill
        R += 1
    put([bundle.get("trip", {}).get("name", "Trip") + " — Budget"], SEC)
    put([f"{_travelers(t)} · currency {cur or '—'} · figures are planning estimates"])
    R += 1
    subs = []
    for sec in bundle.get("budget", {}).get("sections", []):
        put([sec["title"]], SEC); hrow = R; put(["Item", "Quantity", "Unit Cost", "Total", "Notes"], BOLD)
        for c in range(1, 6):
            x = ws.cell(row=hrow, column=c); x.fill = HF; x.font = HFONT; x.alignment = CTR; x.border = BORDER
        start = R
        for it in sec["items"]:
            ws.cell(row=R, column=1, value=it.get("item")).border = BORDER
            q = ws.cell(row=R, column=2, value=it.get("qty")); q.border = BORDER
            u = ws.cell(row=R, column=3, value=it.get("unit")); u.border = BORDER; u.number_format = money
            tot = ws.cell(row=R, column=4); tot.border = BORDER; tot.number_format = money
            if isinstance(it.get("qty"), (int, float)) and isinstance(it.get("unit"), (int, float)):
                tot.value = f"=B{R}*C{R}"
            else:
                tot.value = it.get("total", 0)
            ws.cell(row=R, column=5, value=it.get("note", "")).border = BORDER
            for c in range(1, 6):
                ws.cell(row=R, column=c).font = BASE; ws.cell(row=R, column=c).alignment = WRAP
            R += 1
        sc = ws.cell(row=R, column=1, value=f"{sec['title']} Subtotal"); sc.font = BOLD; sc.fill = SUB; sc.border = BORDER
        for c in (2, 3, 5): ws.cell(row=R, column=c).fill = SUB; ws.cell(row=R, column=c).border = BORDER
        st = ws.cell(row=R, column=4, value=f"=SUM(D{start}:D{R-1})"); st.font = BOLD; st.fill = SUB
        st.number_format = money; st.border = BORDER
        subs.append(R); R += 2
    put(["Grand Total"], SEC); gh = R; put(["Category", "Total"], BOLD)
    for c in (1, 2):
        x = ws.cell(row=gh, column=c); x.fill = HF; x.font = HFONT; x.alignment = CTR; x.border = BORDER
    gstart = R
    for sec, sr in zip(bundle.get("budget", {}).get("sections", []), subs):
        ws.cell(row=R, column=1, value=sec["title"]).border = BORDER
        b = ws.cell(row=R, column=2, value=f"=D{sr}"); b.number_format = money; b.border = BORDER
        ws.cell(row=R, column=1).font = BASE; b.font = BASE; R += 1
    ov = R
    a = ws.cell(row=R, column=1, value="Overall Total"); a.font = BOLD; a.fill = SUB; a.border = BORDER
    b = ws.cell(row=R, column=2, value=f"=SUM(B{gstart}:B{R-1})"); b.font = BOLD; b.fill = SUB
    b.number_format = money; b.border = BORDER; R += 1
    pct = bundle.get("budget", {}).get("contingency_pct", 10)
    a = ws.cell(row=R, column=1, value=f"Contingency ({pct}%)"); a.border = BORDER; a.font = BASE
    b = ws.cell(row=R, column=2, value=f"=B{ov}*{pct/100}"); b.number_format = money; b.border = BORDER; b.font = BASE
    cont = R; R += 1
    a = ws.cell(row=R, column=1, value="Adjusted Grand Total"); a.font = BOLD
    a.fill = PatternFill("solid", fgColor="F2DCDB"); a.border = BORDER
    b = ws.cell(row=R, column=2, value=f"=B{ov}+B{cont}"); b.font = BOLD
    b.fill = PatternFill("solid", fgColor="F2DCDB"); b.number_format = money; b.border = BORDER
    for i, w in enumerate([34, 14, 16, 16, 40], 1): ws.column_dimensions[get_column_letter(i)].width = w

    # 5/6/7 prose sheets
    _prose(wb, "Costs Remarks", bundle.get("costs_remarks", []))
    _prose(wb, "Notes and tips", bundle.get("notes", []))
    _prose(wb, "Prompt", bundle.get("prompt", []))

    wb.save(out)
    return out


def _grid(ws, r0, ncols):
    for rr in range(r0, ws.max_row + 1):
        for cc in range(1, ncols + 1):
            x = ws.cell(row=rr, column=cc); x.border = BORDER
            if cc <= ncols - 3: x.font = BASE; x.alignment = WRAP


def _prose(wb, title, lines):
    ws = wb.create_sheet(title); ws.column_dimensions["A"].width = 110; r = 1
    for ln in lines:
        if isinstance(ln, (list, tuple)) and len(ln) == 2:
            text, kind = ln
        else:
            text, kind = (ln, "n")
        c = ws.cell(row=r, column=1, value=text); c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = SEC if kind == "h" else BASE; r += 1


def _travelers(t):
    tr = t.get("travelers") or {}
    if not tr: return ""
    return f"{tr.get('adults', '?')} adults + {tr.get('children', 0)} children"


def _sym(cur):
    return {"EUR": "€", "USD": "$", "GBP": "£", "ILS": "₪", "NIS": "₪"}.get(cur.upper(), cur + " ")


if __name__ == "__main__":
    bundle = json.load(open(sys.argv[1]))
    out = sys.argv[2] if len(sys.argv) > 2 else "trip.xlsx"
    print(build(bundle, out))
